from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook, Workbook  # Import Workbook to create a new Excel file

app = Flask(__name__)

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    name = request.form.get("name")
    email = request.form.get("email")
    message = request.form.get("message")
    
    # Open the Excel file or create a new one if it doesn't exist
    try:
        workbook = load_workbook("contact_data.xlsx")
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Name"
        sheet["B1"] = "Email"
        sheet["C1"] = "Message"
    
    sheet = workbook.active
    
    # Find the next available row
    next_row = sheet.max_row + 1
    
    # Write the data to the Excel file
    sheet[f"A{next_row}"] = name
    sheet[f"B{next_row}"] = email
    sheet[f"C{next_row}"] = message
    
    # Save the file
    workbook.save("contact_data.xlsx")
    
    # Redirect to the Thank You page
    return redirect(url_for("thank_you"))

@app.route("/thank-you")
def thank_you():
    return render_template("thank_you.html")
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
